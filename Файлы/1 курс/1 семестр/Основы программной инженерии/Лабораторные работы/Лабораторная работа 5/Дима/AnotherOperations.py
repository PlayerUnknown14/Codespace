import time
from tkinter import *
from tkinter import ttk
import tkinter as tk
from math import sqrt, log10
import numpy as np
import openpyxl

wb = openpyxl.load_workbook("Results.xlsx")
wb.create_sheet('Sum')
wb.create_sheet('Razn')
wb.save("Results.xlsx")
ws = wb['Sum']
root = tk.Tk()
root.title('AnotherMathOperations')
root.geometry('500x500')
massgen = []

def genoption():
    match Comboboxtype.get().split():
        case ['Int']:
            Gen.integer()
        case ['Float']:
            Gen.float()
def clicked():
    global ws
    match Comboboxtype.get().split():
        case ["Int"]:
            match Comboboxop.get().split():
                case ["Сложение"]:
                    howmuch = int(howmuch_entry.get())
                    
                    for i in range(1, howmuch+1):
                        typeofgen, timeofgen, lengthofgen, rangeofgen, summass = IntOperations.sum()
                        wb = openpyxl.load_workbook("Results.xlsx")
                        ws = wb['Sum']
                        for j in range(1,howmuch+1):
                            ws.cell(row = i+1, column=j+1) == typeofgen                
                            ws.cell(row = i+2, column=j+1) == timeofgen  
                            ws.cell(row = i+3, column=j+1) == lengthofgen  
                            ws.cell(row = i+4, column=j+1) == summass 
                            ws.cell(row = i+5, column=j+1) == rangeofgen

            
        
class Gen:
    def integer():
        global timeofgen, massgen
        start_time = time.time()
        
        amount = amount_entry.get()
        rangeran = range_entry.get().split(' ')
        rangemin = int(rangeran[0])
        rangemax = int(rangeran[1])
        massgen = np.random.randint(rangemin, rangemax, int(amount))
        
        end_time = time.time()
        timeofgen = end_time - start_time
    def float():
        global timeofgen, massgen
        start_time = time.time()
        
        amount = amount_entry.get()
        rangeran = range_entry.get().split(' ')
        rangemin = float(rangeran[0])
        rangemax = float(rangeran[1])
        massgen = np.random.uniform(rangemin, rangemax, int(amount))
        
        end_time = time.time()
        timeofgen = end_time - start_time
        
class IntOperations:
    def sum():
        global massgen
        summass = 0
        start_time = time.time()
        for i in massgen:
            summass += i
        end_time = time.time()
        timeofgen = end_time - start_time
        lengthofgen = len(massgen)
        rangeofgen = f'{min(massgen)};{max(massgen)}'
        match Comboboxtype.get().split():
            case['Int']:
                typeofgen = 'Int Sum'
            case['Float']:
                typeofgen = 'Float Sum'
        return(typeofgen, timeofgen, lengthofgen, rangeofgen, summass)
 

Operationslist = ['Сложение', 'Вычитание', 'Умножение', 'Деление', 'Корни', 'Логарифм']
Operations_var = StringVar(value=Operationslist[0])

Comboboxop = ttk.Combobox(textvariable=Operations_var, values=Operationslist, state='readonly')
Comboboxop.place(x=6,y=6)

Typelist = ['Int', 'Float']
Typelist_var = StringVar(value=Typelist[0])

Comboboxtype = ttk.Combobox(textvariable=Typelist_var, values=Typelist, state='readonly')
Comboboxtype.place(x=156, y=6)


num = 30

amount_label = ttk.Label(text='Введите количество элементов массива')
amount_label.place(x=6,y=num)

amount_entry = ttk.Entry()
amount_entry.place(x=6, y=num+20)

range_label = ttk.Label(text='Введите через пробел диапазон генерируемых чисел')
range_label.place(x=6, y =num*2+20)

range_entry = ttk.Entry()
range_entry.place(x=6, y=num*3+20)

gen_button = ttk.Button(text='Сгенерировать массив', command=genoption)

start_button = ttk.Button(text='Старт', command=clicked)
start_button.place(x=6, y=num*4+20)

time_label = ttk.Label(text='')
time_label.place(x=6, y=num*5+20)

timemass_label = ttk.Label(text='')
timemass_label.place(x=6, y=num*6+20)

mass_label = ttk.Label(text='')
mass_label.place(x=6, y=num*7+20)

howmuch_entry = ttk.Entry()
howmuch_entry.place(x=350, y=6)

root.mainloop()
